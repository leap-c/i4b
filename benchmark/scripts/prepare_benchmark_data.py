"""Download and normalize source data for the multi-building benchmark.

Acquires TABULA building records, Open-Meteo reference weather and archived
forecast runs, and Energy-Charts day-ahead electricity prices. Normalizes
each into typed Parquet tables with checksummed provenance manifests. This
script only prepares source data; it does not touch I4B simulation code and
is never imported by it. See tickets/source-data-acquisition.md.

Usage:
    python scripts/prepare_benchmark_data.py \\
        --config scripts/benchmark_source_data.json \\
        --output-dir /path/to/source-data
"""

from __future__ import annotations

import argparse
import hashlib
import itertools
import json
import sys
import time
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

CONFIG_SCHEMA_VERSION = 1


class ConfigError(ValueError):
    """Raised when the acquisition configuration fails validation."""


@dataclass(frozen=True)
class Period:
    id: str
    start: date
    end: date


@dataclass(frozen=True)
class Location:
    id: str
    country: str
    latitude: float
    longitude: float
    timezone: str
    market: str | None
    altitude: float = 0.0


@dataclass(frozen=True)
class BenchmarkConfig:
    periods: list[Period]
    weather_model: str
    forecast_run_hours_utc: list[int]
    forecast_horizon_hours: int
    locations: list[Location]


def load_config(path: Path) -> BenchmarkConfig:
    return parse_config(json.loads(Path(path).read_text()))


def parse_config(raw: dict[str, Any]) -> BenchmarkConfig:
    if raw.get("schema_version") != CONFIG_SCHEMA_VERSION:
        raise ConfigError(
            f"Unsupported schema_version: {raw.get('schema_version')!r} "
            f"(expected {CONFIG_SCHEMA_VERSION})"
        )

    periods = [_parse_period(p) for p in raw.get("periods", [])]
    if not periods:
        raise ConfigError("Configuration must declare at least one period")
    _require_unique([p.id for p in periods], "period id")
    for period in periods:
        if period.start >= period.end:
            raise ConfigError(f"Period '{period.id}' start must be before end")

    weather_model = raw.get("weather_model")
    if not isinstance(weather_model, str) or not weather_model:
        raise ConfigError("Configuration must declare a non-empty 'weather_model'")

    run_hours = raw.get("forecast_run_hours_utc", [])
    if not run_hours or not all(isinstance(h, int) and 0 <= h <= 23 for h in run_hours):
        raise ConfigError(
            "'forecast_run_hours_utc' must be a non-empty list of hours 0-23"
        )

    horizon = raw.get("forecast_horizon_hours")
    if not isinstance(horizon, int) or horizon <= 0:
        raise ConfigError("'forecast_horizon_hours' must be a positive integer")

    locations = [_parse_location(loc) for loc in raw.get("locations", [])]
    if not locations:
        raise ConfigError("Configuration must declare at least one location")
    _require_unique([loc.id for loc in locations], "location id")
    _require_unique([loc.country for loc in locations], "location country")
    _require_unique([loc.market for loc in locations if loc.market], "location market")

    return BenchmarkConfig(
        periods=periods,
        weather_model=weather_model,
        forecast_run_hours_utc=sorted(run_hours),
        forecast_horizon_hours=horizon,
        locations=locations,
    )


def _parse_period(raw: dict[str, Any]) -> Period:
    try:
        return Period(
            id=raw["id"],
            start=date.fromisoformat(raw["start"]),
            end=date.fromisoformat(raw["end"]),
        )
    except (KeyError, ValueError) as exc:
        raise ConfigError(f"Invalid period entry {raw!r}: {exc}") from exc


def _parse_location(raw: dict[str, Any]) -> Location:
    try:
        return Location(
            id=raw["id"],
            country=raw["country"],
            latitude=float(raw["latitude"]),
            longitude=float(raw["longitude"]),
            timezone=raw["timezone"],
            market=raw.get("market"),
            altitude=float(raw.get("altitude", 0.0)),
        )
    except (KeyError, ValueError) as exc:
        raise ConfigError(f"Invalid location entry {raw!r}: {exc}") from exc


def _require_unique(values: list[str], label: str) -> None:
    duplicates = sorted({v for v in values if values.count(v) > 1})
    if duplicates:
        raise ConfigError(f"Duplicate {label} values: {duplicates}")


# ---------------------------------------------------------------------------
# Storage and provenance
# ---------------------------------------------------------------------------


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def write_bytes_atomic(data: bytes, path: Path) -> str:
    """Write raw bytes via a temp file, returning their checksum."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + ".tmp")
    tmp_path.write_bytes(data)
    checksum = sha256_file(tmp_path)
    tmp_path.replace(path)
    return checksum


def write_dataframe_atomic(df: pd.DataFrame, path: Path) -> str:
    """Write ``df`` to Parquet via a temp file, returning the file's checksum."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + ".tmp")
    df.to_parquet(tmp_path, engine="pyarrow", index=False)
    checksum = sha256_file(tmp_path)
    tmp_path.replace(path)
    return checksum


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + ".tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, default=str) + "\n")
    tmp_path.replace(path)


def null_counts(df: pd.DataFrame) -> dict[str, int]:
    """Per-column count of missing values, omitting columns with none.

    Lets a later loader tell an expected source gap (e.g. null radiation at
    forecast lead zero) from a corrupted artifact.
    """
    return {str(col): int(n) for col, n in df.isna().sum().items() if n > 0}


def build_manifest(
    *,
    artifact_type: str,
    source_url: str,
    request: dict[str, Any],
    retrieved_at_utc: str,
    raw_sha256: str,
    normalized_sha256: str,
    df: pd.DataFrame,
    column_units: dict[str, str],
    license_note: str,
    time_column: str | None = None,
    raw_request: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Assemble the provenance sidecar written next to every normalized table.

    ``raw_request`` is recorded only when the HTTP call covered more than this
    artifact - weather is fetched for all locations at once, so the manifest
    would otherwise not describe the request that actually produced the file.
    """
    manifest = {
        "schema_version": CONFIG_SCHEMA_VERSION,
        "artifact_type": artifact_type,
        "source_url": source_url,
        "request": request,
        "retrieved_at_utc": retrieved_at_utc,
        "raw_sha256": raw_sha256,
        "normalized_sha256": normalized_sha256,
        "row_count": len(df),
        "time_range": None,
        "column_units": column_units,
        "null_counts": null_counts(df),
        "license_note": license_note,
    }
    if time_column:
        manifest["time_range"] = {
            "start": df[time_column].min().isoformat(),
            "end": df[time_column].max().isoformat(),
        }
    if raw_request is not None and raw_request != request:
        manifest["raw_request"] = raw_request
    return manifest


def artifact_up_to_date(
    parquet_path: Path, manifest_path: Path, request: dict[str, Any] | None = None
) -> bool:
    """True if the artifact exists, matches its manifest checksum, and - when
    ``request`` is given - was produced by that same request.

    The request comparison is what stops an edited configuration (moved
    coordinates, shifted dates) from silently reusing and relabelling data
    fetched for something else.
    """
    if not parquet_path.exists() or not manifest_path.exists():
        return False
    try:
        manifest = json.loads(manifest_path.read_text())
    except json.JSONDecodeError:
        return False
    if manifest.get("normalized_sha256") != sha256_file(parquet_path):
        return False
    return request is None or manifest.get("request") == request


def _raw_sidecar_path(raw_path: Path) -> Path:
    return raw_path.with_name(raw_path.stem + ".request.json")


def read_raw_cache(
    raw_path: Path, request: dict[str, Any], *, force: bool
) -> str | None:
    """Return the original retrieval timestamp if ``raw_path`` caches exactly
    ``request``, else ``None`` so the caller refetches.

    Returning the original timestamp keeps ``retrieved_at_utc`` honest when a
    normalized table is rebuilt from a cached response.
    """
    if force:
        return None
    sidecar_path = _raw_sidecar_path(raw_path)
    if not raw_path.exists() or not sidecar_path.exists():
        return None
    try:
        sidecar = json.loads(sidecar_path.read_text())
    except json.JSONDecodeError:
        return None
    if sidecar.get("request") != request:
        print(f"[cache] {raw_path.name}: request changed, refetching")
        return None
    return sidecar.get("retrieved_at_utc")


def write_raw_cache(
    raw_path: Path, *, source_url: str, request: dict[str, Any], retrieved_at_utc: str
) -> None:
    write_json(
        _raw_sidecar_path(raw_path),
        {
            "source_url": source_url,
            "request": request,
            "retrieved_at_utc": retrieved_at_utc,
        },
    )


def _to_float(value: Any) -> float | None:
    return None if value is None or value == "" else float(value)


def _to_int(value: Any) -> int | None:
    return None if value is None or value == "" else int(value)


def _to_str(value: Any) -> str | None:
    return None if value is None else str(value)


RETRY_STATUS = {429, 500, 502, 503, 504}


def _get(url: str, params: dict[str, Any] | None = None, *, attempts: int = 5):
    """GET with backoff on rate limits and transient server errors.

    Both providers throttle: Open-Meteo weighs a batched request once per
    location, so a few hundred runs exhaust the per-minute allowance, and
    Energy-Charts rejects even a handful of back-to-back calls. Callers keep
    their own handling of the final response.
    """
    import requests

    session = requests.Session()
    for attempt in range(attempts):
        try:
            response = session.get(url, params=params, timeout=120)
        except requests.RequestException as exc:
            if attempt == attempts - 1:
                raise
            delay = 5 * 2**attempt
            print(f"[retry] {type(exc).__name__}; waiting {delay}s")
            time.sleep(delay)
            continue
        if response.status_code not in RETRY_STATUS or attempt == attempts - 1:
            return response
        delay = 5 * 2**attempt
        print(f"[retry] HTTP {response.status_code}; waiting {delay}s")
        time.sleep(delay)
    return response


# ---------------------------------------------------------------------------
# TABULA buildings
# ---------------------------------------------------------------------------

TABULA_URL = "https://episcope.eu/fileadmin/tabula/public/calc/tabula-calculator.xlsx"
# The published workbook keeps the building-variant catalogue on
# "Calc.Set.Building"; "Calc.Building" (as named in early ticket drafts) does
# not exist. Data starts below a template row, so rows are selected by the
# filter below rather than by a fixed row offset.
TABULA_SHEET = "Calc.Set.Building"

TABULA_COUNTRIES = {"BG", "CY", "DE", "FR", "IE", "NO", "PL"}
TABULA_SIZE_CLASS = "SFH"
TABULA_DATA_TYPE = "ReEx"
TABULA_DATASET_STATUS = "Typology"
TABULA_VARIANTS = {1, 2, 3}

_CONVERTERS = {"string": _to_str, "Int64": _to_int, "Float64": _to_float}

# normalized column | TABULA source column | dtype | unit
TABULA_FIELDS: list[tuple[str, str, str, str]] = [
    ("building_id", "Code_BuildingVariant", "string", "string"),
    ("building_family_id", "Code_Building", "string", "string"),
    ("country_code", "Code_Country", "string", "ISO 3166-1 alpha-2"),
    (
        "variant_number",
        "Number_BuildingVariant",
        "Int64",
        "1=existing 2=standard 3=ambitious",
    ),
    ("dataset_status", "Code_StatusDataset", "string", "string"),
    ("data_type", "Code_DataType_Building", "string", "string"),
    ("building_size_class", "Code_BuildingSizeClass", "string", "string"),
    ("year_start", "Year1_Building", "Int64", "year"),
    ("year_end", "Year2_Building", "Int64", "year"),
    ("roof_type", "Code_RoofType", "string", "string"),
    ("attached_neighbours", "Code_AttachedNeighbours", "string", "string"),
    ("reference_area_m2", "A_C_Ref", "Float64", "m^2"),
    ("room_height_m", "h_room", "Float64", "m"),
    ("thermal_capacity_Wh_m2K", "c_m", "Float64", "Wh/(m^2 K)"),
    ("transmission_W_m2K", "h_Transmission", "Float64", "W/(m^2 K)"),
    ("ventilation_W_m2K", "h_Ventilation", "Float64", "W/(m^2 K)"),
    ("window_1_area_m2", "A_Window_1", "Float64", "m^2"),
    ("window_2_area_m2", "A_Window_2", "Float64", "m^2"),
    ("window_east_area_m2", "A_Window_East", "Float64", "m^2"),
    ("window_south_area_m2", "A_Window_South", "Float64", "m^2"),
    ("window_west_area_m2", "A_Window_West", "Float64", "m^2"),
    ("window_north_area_m2", "A_Window_North", "Float64", "m^2"),
    ("window_g_value", "g_gl_n", "Float64", "dimensionless"),
    ("window_1_transmission_W_K", "H_Transmission_Window_1", "Float64", "W/K"),
    ("window_2_transmission_W_K", "H_Transmission_Window_2", "Float64", "W/K"),
    ("door_1_transmission_W_K", "H_Transmission_Door_1", "Float64", "W/K"),
]

TABULA_SOURCE_COLUMNS = [source for _, source, _, _ in TABULA_FIELDS]
TABULA_COLUMN_UNITS = {name: unit for name, _, _, unit in TABULA_FIELDS} | {
    "window_orientation_missing": "boolean"
}

TABULA_LICENSE_NOTE = (
    "Source: EPISCOPE/TABULA building typology database (episcope.eu). "
    "Review episcope.eu terms of use before redistribution."
)

# Cohort shape of the currently published workbook: country -> (variants, families).
# Totals are 191 variants across 64 families. An upstream revision that adds,
# removes, or reclassifies variants must fail loudly rather than silently
# change the benchmark population.
TABULA_EXPECTED_COHORT = {
    "BG": (21, 7),
    "CY": (12, 4),
    "DE": (39, 13),
    "FR": (30, 10),
    "IE": (44, 15),
    "NO": (24, 8),
    "PL": (21, 7),
}
# Norwegian variants carrying window area with no cardinal orientation split.
TABULA_EXPECTED_ORIENTATION_MISSING = 21


def download_workbook(url: str, destination: Path) -> str:
    """Download the TABULA workbook, returning its checksum."""
    response = _get(url)
    response.raise_for_status()
    return write_bytes_atomic(response.content, destination)


def read_building_rows(workbook_path: Path) -> list[dict[str, Any]]:
    """Read raw building-variant rows from the workbook's data sheet.

    ``data_only=True`` is required: TABULA stores formula cells and the
    normalizer needs their cached values, not the formula text.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows_iter = workbook[TABULA_SHEET].iter_rows(values_only=True)
        header = next(rows_iter)
        header_index = {
            name: idx for idx, name in enumerate(header) if name is not None
        }

        missing = set(TABULA_SOURCE_COLUMNS) - header_index.keys()
        if missing:
            raise ValueError(
                f"TABULA workbook missing expected columns: {sorted(missing)}"
            )

        return [
            {name: row[idx] for name, idx in header_index.items()}
            for row in rows_iter
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()


def select_sfh_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Filter raw TABULA rows to the ReEx SFH Typology variants 1-3."""
    selected = []
    for row in rows:
        if (
            row.get("Code_StatusDataset") != TABULA_DATASET_STATUS
            or row.get("Code_Country") not in TABULA_COUNTRIES
            or row.get("Code_BuildingSizeClass") != TABULA_SIZE_CLASS
            or row.get("Code_DataType_Building") != TABULA_DATA_TYPE
        ):
            continue
        try:
            variant = int(row["Number_BuildingVariant"])
        except (KeyError, TypeError, ValueError):
            continue
        if variant in TABULA_VARIANTS:
            selected.append(row)
    return selected


def _orientation_missing(record: dict[str, Any]) -> bool:
    """True when a building has window area but no cardinal orientation split.

    Flags the 21 Norwegian variants. The ticket forbids imputing those areas
    during normalization, so they are marked and left alone.
    """
    total = sum(record[f"window_{n}_area_m2"] or 0.0 for n in (1, 2))
    cardinal = sum(
        record[f"window_{d}_area_m2"] or 0.0 for d in ("east", "south", "west", "north")
    )
    return total > 0 and cardinal == 0


def normalize_building_rows(rows: list[dict[str, Any]]) -> pd.DataFrame:
    """Normalize selected TABULA rows into the flat building table."""
    records = []
    for row in rows:
        record = {
            name: _CONVERTERS[dtype](row.get(source))
            for name, source, dtype, _ in TABULA_FIELDS
        }
        record["window_orientation_missing"] = _orientation_missing(record)
        records.append(record)

    if not records:
        raise ValueError("No TABULA rows survived normalization")

    df = pd.DataFrame.from_records(records)
    df = df.sort_values(["country_code", "building_family_id", "variant_number"])
    df = df.reset_index(drop=True)

    for name, _, dtype, _ in TABULA_FIELDS:
        df[name] = df[name].astype(dtype)
    df["window_orientation_missing"] = df["window_orientation_missing"].astype(
        "boolean"
    )
    return df


def validate_tabula_cohort(
    df: pd.DataFrame,
    expected_cohort: dict[str, tuple[int, int]] = TABULA_EXPECTED_COHORT,
    expected_orientation_missing: int = TABULA_EXPECTED_ORIENTATION_MISSING,
) -> None:
    """Check the normalized table against the known population of the workbook.

    ``normalize_building_rows`` stays a general transformer testable with a
    handful of rows; this asserts the cohort it should produce from the real
    published workbook.
    """
    errors = []

    if df["building_id"].isna().any() or df["building_id"].duplicated().any():
        errors.append("building_id must be unique and non-null")

    actual = {
        country: (len(group), group["building_family_id"].nunique())
        for country, group in df.groupby("country_code", observed=True)
    }
    if actual != dict(expected_cohort):
        errors.append(
            f"(variants, families) per country {actual} != {dict(expected_cohort)}"
        )

    n_missing = int(df["window_orientation_missing"].fillna(False).astype(bool).sum())
    if n_missing != expected_orientation_missing:
        errors.append(
            f"expected {expected_orientation_missing} orientation-missing variants, "
            f"got {n_missing}"
        )

    if errors:
        raise ValueError("TABULA cohort validation failed: " + "; ".join(errors))


def run_tabula(output_dir: Path, *, force: bool) -> None:
    raw_path = output_dir / "raw" / "tabula" / "tabula-calculator.xlsx"
    normalized_path = output_dir / "normalized" / "buildings" / "tabula_sfh.parquet"
    manifest_path = output_dir / "manifests" / "buildings" / "tabula_sfh.json"

    request = {
        "sheet": TABULA_SHEET,
        "Code_StatusDataset": TABULA_DATASET_STATUS,
        "Code_Country": sorted(TABULA_COUNTRIES),
        "Code_BuildingSizeClass": TABULA_SIZE_CLASS,
        "Code_DataType_Building": TABULA_DATA_TYPE,
        "Number_BuildingVariant": sorted(TABULA_VARIANTS),
    }

    if not force and artifact_up_to_date(normalized_path, manifest_path, request):
        print(f"[tabula] up to date: {normalized_path}")
        return

    retrieved_at_utc = read_raw_cache(raw_path, request, force=force)
    if retrieved_at_utc:
        print(f"[tabula] reusing cached workbook: {raw_path}")
        raw_sha256 = sha256_file(raw_path)
    else:
        print(f"[tabula] downloading {TABULA_URL}")
        raw_sha256 = download_workbook(TABULA_URL, raw_path)
        retrieved_at_utc = utc_now_iso()
        write_raw_cache(
            raw_path,
            source_url=TABULA_URL,
            request=request,
            retrieved_at_utc=retrieved_at_utc,
        )

    df = normalize_building_rows(select_sfh_rows(read_building_rows(raw_path)))
    validate_tabula_cohort(df)
    print(
        f"[tabula] {len(df)} variants across {df['building_family_id'].nunique()} families"
    )

    normalized_sha256 = write_dataframe_atomic(df, normalized_path)
    write_json(
        manifest_path,
        build_manifest(
            artifact_type="tabula_buildings",
            source_url=TABULA_URL,
            request=request,
            retrieved_at_utc=retrieved_at_utc,
            raw_sha256=raw_sha256,
            normalized_sha256=normalized_sha256,
            df=df,
            column_units=TABULA_COLUMN_UNITS,
            license_note=TABULA_LICENSE_NOTE,
        ),
    )
    print(f"[tabula] wrote {normalized_path}")


# ---------------------------------------------------------------------------
# Weather (Open-Meteo)
# ---------------------------------------------------------------------------

OPEN_METEO_ARCHIVE_URL = "https://archive-api.open-meteo.com/v1/archive"
OPEN_METEO_SINGLE_RUNS_URL = "https://single-runs-api.open-meteo.com/v1/forecast"

WEATHER_HOURLY_VARIABLES = [
    "temperature_2m",
    "shortwave_radiation",
    "direct_normal_irradiance",
    "diffuse_radiation",
]

# The archive endpoint serves a blended reanalysis and ignores a `models=`
# selector, so reference weather is labelled for what it is. `weather_model`
# from the configuration applies to forecast runs only.
OPEN_METEO_REFERENCE_MODEL = "open_meteo_archive"

# Individual ECMWF IFS HRES runs are documented from this date; an earlier
# request must fail rather than silently fall back to another model.
ECMWF_IFS_AVAILABLE_FROM = date(2024, 3, 14)

OPEN_METEO_LICENSE_NOTE = (
    "Open-Meteo weather data, CC BY 4.0 (open-meteo.com/en/license)."
)

# These also fix the normalized column order.
WEATHER_REFERENCE_COLUMN_UNITS = {
    "location_id": "string",
    "valid_time_utc": "timestamp[UTC]",
    "model": "string",
    "temperature_2m_C": "degC",
    "ghi_W_m2": "W/m^2",
    "dni_W_m2": "W/m^2",
    "dhi_W_m2": "W/m^2",
}
WEATHER_FORECAST_COLUMN_UNITS = {
    "location_id": "string",
    "model": "string",
    "initialization_time_utc": "timestamp[UTC]",
    "valid_time_utc": "timestamp[UTC]",
    "lead_hours": "hours",
    "temperature_2m_C": "degC",
    "ghi_W_m2": "W/m^2",
    "dni_W_m2": "W/m^2",
    "dhi_W_m2": "W/m^2",
}


def _location_entry(location: Location) -> dict[str, Any]:
    return {
        "location_id": location.id,
        "latitude": location.latitude,
        "longitude": location.longitude,
    }


def narrow_request(request: dict[str, Any], location: Location) -> dict[str, Any]:
    """The batched request as it applies to one location's artifact."""
    return {**request, "locations": [_location_entry(location)]}


def reference_request(locations: list[Location], period: Period) -> dict[str, Any]:
    return {
        "locations": [_location_entry(loc) for loc in locations],
        "start_date": period.start.isoformat(),
        "end_date": period.end.isoformat(),
        "hourly": WEATHER_HOURLY_VARIABLES,
        "timezone": "GMT",
    }


def forecast_request(
    locations: list[Location],
    model: str,
    initialization_time: datetime,
    horizon_hours: int,
) -> dict[str, Any]:
    return {
        "locations": [_location_entry(loc) for loc in locations],
        "model": model,
        "run": initialization_time.isoformat(),
        "forecast_hours": horizon_hours,
        "hourly": WEATHER_HOURLY_VARIABLES,
        "timezone": "GMT",
    }


def _open_meteo_params(locations: list[Location]) -> dict[str, str]:
    """Open-Meteo takes comma-separated coordinates, so all configured
    locations are fetched in one call per period or run."""
    return {
        "latitude": ",".join(str(loc.latitude) for loc in locations),
        "longitude": ",".join(str(loc.longitude) for loc in locations),
        "hourly": ",".join(WEATHER_HOURLY_VARIABLES),
        "timezone": "GMT",
    }


def _as_location_list(payload: Any) -> list[dict[str, Any]]:
    return payload if isinstance(payload, list) else [payload]


def download_reference_weather(
    locations: list[Location], start: date, end: date
) -> list[dict[str, Any]]:
    """Download batched archive weather for all ``locations`` over one period."""
    params = {
        **_open_meteo_params(locations),
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
    }
    response = _get(OPEN_METEO_ARCHIVE_URL, params)
    response.raise_for_status()
    return _as_location_list(response.json())


def download_forecast_run(
    locations: list[Location],
    model: str,
    initialization_time: datetime,
    horizon_hours: int,
) -> list[dict[str, Any]]:
    """Download one batched archived forecast run for all ``locations``."""
    if model == "ecmwf_ifs" and initialization_time.date() < ECMWF_IFS_AVAILABLE_FROM:
        raise ValueError(
            f"ECMWF IFS runs are only documented from {ECMWF_IFS_AVAILABLE_FROM.isoformat()}; "
            f"requested run {initialization_time.isoformat()} is unavailable."
        )

    params = {
        **_open_meteo_params(locations),
        "models": model,
        "run": initialization_time.strftime("%Y-%m-%dT%H:%M"),
        "forecast_hours": horizon_hours,
    }
    response = _get(OPEN_METEO_SINGLE_RUNS_URL, params)
    if response.status_code == 400:
        if "requested model run is not available" in response.text:
            raise LookupError(response.text)
        raise ValueError(f"Open-Meteo rejected forecast run request: {response.text}")
    response.raise_for_status()
    return _as_location_list(response.json())


def normalize_weather_response(
    payload: list[dict[str, Any]],
    locations: list[Location],
    *,
    model: str,
    initialization_time: datetime | None = None,
) -> dict[str, pd.DataFrame]:
    """Split a batched Open-Meteo response into one frame per location.

    ``initialization_time=None`` produces reference-weather columns; a
    timestamp produces forecast-run columns including ``lead_hours``.
    """
    if len(payload) != len(locations):
        raise ValueError(
            f"Open-Meteo response has {len(payload)} locations, expected {len(locations)}"
        )

    frames = {}
    for location, location_payload in zip(locations, payload):
        hourly = location_payload["hourly"]
        valid_time = pd.to_datetime(hourly["time"], utc=True)

        data: dict[str, Any] = {
            "location_id": location.id,
            "valid_time_utc": valid_time,
            "temperature_2m_C": pd.array(hourly["temperature_2m"], dtype="Float64"),
            "ghi_W_m2": pd.array(hourly["shortwave_radiation"], dtype="Float64"),
            "dni_W_m2": pd.array(hourly["direct_normal_irradiance"], dtype="Float64"),
            "dhi_W_m2": pd.array(hourly["diffuse_radiation"], dtype="Float64"),
        }

        if initialization_time is None:
            data["model"] = OPEN_METEO_REFERENCE_MODEL
            columns = list(WEATHER_REFERENCE_COLUMN_UNITS)
        else:
            init_ts = pd.Timestamp(initialization_time)
            if init_ts.tzinfo is None:
                init_ts = init_ts.tz_localize("UTC")
            data["model"] = model
            data["initialization_time_utc"] = init_ts
            data["lead_hours"] = (
                (valid_time - init_ts).total_seconds() / 3600.0
            ).astype("float64")
            columns = list(WEATHER_FORECAST_COLUMN_UNITS)

        df = pd.DataFrame(data)[columns]
        times = df["valid_time_utc"]
        if times.duplicated().any() or not times.is_monotonic_increasing:
            raise ValueError(
                f"Non-increasing or duplicate valid_time_utc for {location.id}"
            )
        frames[location.id] = df

    return frames


def iter_configured_forecast_runs(
    config: BenchmarkConfig,
) -> Iterator[tuple[Period, datetime]]:
    """Deterministically enumerate (period, initialization_time_utc) runs."""
    for period in config.periods:
        day = period.start
        while day <= period.end:
            for hour in config.forecast_run_hours_utc:
                yield (
                    period,
                    datetime(day.year, day.month, day.day, hour, tzinfo=timezone.utc),
                )
            day += timedelta(days=1)


def check_forecast_availability(config: BenchmarkConfig) -> None:
    """Probe the earliest run of every configured hour before a long acquisition.

    Availability varies by run hour, not just by date: Open-Meteo's ECMWF IFS
    archive carries 00Z/12Z from the start but only serves 06Z/18Z from around
    2024-08-09. Probing just the first run would pass on 00Z and then fail
    hours into the job, so each configured hour is checked in each period.
    """
    seen = set()
    for period, run in iter_configured_forecast_runs(config):
        key = (period.id, run.hour)
        if key in seen:
            continue
        seen.add(key)
        print(f"[weather-forecast] preflight check: run {run.isoformat()}")
        payload = download_forecast_run(
            config.locations, config.weather_model, run, horizon_hours=1
        )
        normalize_weather_response(
            payload,
            config.locations,
            model=config.weather_model,
            initialization_time=run,
        )


def run_weather_reference(
    config: BenchmarkConfig, output_dir: Path, *, force: bool
) -> None:
    for period in config.periods:
        request = reference_request(config.locations, period)
        raw_path = output_dir / "raw" / "open_meteo" / f"reference_{period.id}.json"

        retrieved_at_utc = read_raw_cache(raw_path, request, force=force)
        if retrieved_at_utc:
            print(f"[weather-reference] reusing cached response: {raw_path}")
            payload = json.loads(raw_path.read_text())
            raw_sha256 = sha256_file(raw_path)
        else:
            print(f"[weather-reference] downloading period {period.id}")
            payload = download_reference_weather(
                config.locations, period.start, period.end
            )
            retrieved_at_utc = utc_now_iso()
            raw_sha256 = write_bytes_atomic(
                json.dumps(payload).encode("utf-8"), raw_path
            )
            write_raw_cache(
                raw_path,
                source_url=OPEN_METEO_ARCHIVE_URL,
                request=request,
                retrieved_at_utc=retrieved_at_utc,
            )

        frames = normalize_weather_response(
            payload, config.locations, model=config.weather_model
        )

        for location in config.locations:
            name = f"{location.id}_{period.id}"
            normalized_path = (
                output_dir / "normalized" / "weather_reference" / f"{name}.parquet"
            )
            manifest_path = (
                output_dir / "manifests" / "weather_reference" / f"{name}.json"
            )
            location_request = narrow_request(request, location)

            if not force and artifact_up_to_date(
                normalized_path, manifest_path, location_request
            ):
                print(f"[weather-reference] up to date: {normalized_path}")
                continue

            df = frames[location.id]
            write_json(
                manifest_path,
                build_manifest(
                    artifact_type="weather_reference",
                    source_url=OPEN_METEO_ARCHIVE_URL,
                    request=location_request,
                    raw_request=request,
                    retrieved_at_utc=retrieved_at_utc,
                    raw_sha256=raw_sha256,
                    normalized_sha256=write_dataframe_atomic(df, normalized_path),
                    df=df,
                    column_units=WEATHER_REFERENCE_COLUMN_UNITS,
                    license_note=OPEN_METEO_LICENSE_NOTE,
                    time_column="valid_time_utc",
                ),
            )
            print(f"[weather-reference] wrote {normalized_path}")


def run_weather_forecasts(
    config: BenchmarkConfig,
    output_dir: Path,
    *,
    force: bool,
    max_runs: int | None = None,
    shard_index: int = 0,
    shard_count: int = 1,
) -> None:
    preflight_done = False
    runs = iter_configured_forecast_runs(config)
    if shard_count != 1 or shard_index != 0:
        runs = itertools.islice(runs, shard_index, None, shard_count)
        print(f"[weather-forecast] shard {shard_index + 1}/{shard_count}")
    if max_runs is not None:
        runs = itertools.islice(runs, max_runs)
        print(f"[weather-forecast] limited to the first {max_runs} runs")

    for period, initialization_time in runs:
        run_id = initialization_time.strftime("%Y%m%dT%H")
        name = f"{config.weather_model}_{run_id}"
        raw_path = output_dir / "raw" / "open_meteo" / f"forecast_{name}.json"
        request = forecast_request(
            config.locations,
            config.weather_model,
            initialization_time,
            config.forecast_horizon_hours,
        )

        def paths(location: Location, name: str = name) -> tuple[Path, Path]:
            return (
                output_dir
                / "normalized"
                / "weather_forecasts"
                / location.id
                / f"{name}.parquet",
                output_dir
                / "manifests"
                / "weather_forecasts"
                / location.id
                / f"{name}.json",
            )

        if not force and all(
            artifact_up_to_date(*paths(loc), narrow_request(request, loc))
            for loc in config.locations
        ):
            print(f"[weather-forecast] up to date: run {run_id}")
            continue

        # Only probe availability once a run in this config actually needs a
        # live request, so a completed acquisition stays reviewable offline.
        if not preflight_done:
            check_forecast_availability(config)
            preflight_done = True

        retrieved_at_utc = read_raw_cache(raw_path, request, force=force)
        if retrieved_at_utc:
            print(f"[weather-forecast] reusing cached response: {raw_path}")
            payload = json.loads(raw_path.read_text())
            raw_sha256 = sha256_file(raw_path)
        else:
            print(f"[weather-forecast] downloading run {run_id} ({period.id})")
            try:
                payload = download_forecast_run(
                    config.locations,
                    config.weather_model,
                    initialization_time,
                    config.forecast_horizon_hours,
                )
            except LookupError as exc:
                print(
                    f"[weather-forecast] unavailable run {run_id}; skipping: {exc}",
                    flush=True,
                )
                continue
            retrieved_at_utc = utc_now_iso()
            raw_sha256 = write_bytes_atomic(
                json.dumps(payload).encode("utf-8"), raw_path
            )
            write_raw_cache(
                raw_path,
                source_url=OPEN_METEO_SINGLE_RUNS_URL,
                request=request,
                retrieved_at_utc=retrieved_at_utc,
            )

        frames = normalize_weather_response(
            payload,
            config.locations,
            model=config.weather_model,
            initialization_time=initialization_time,
        )

        for location in config.locations:
            normalized_path, manifest_path = paths(location)
            location_request = narrow_request(request, location)
            if not force and artifact_up_to_date(
                normalized_path, manifest_path, location_request
            ):
                continue

            df = frames[location.id]
            write_json(
                manifest_path,
                build_manifest(
                    artifact_type="weather_forecast_run",
                    source_url=OPEN_METEO_SINGLE_RUNS_URL,
                    request=location_request,
                    raw_request=request,
                    retrieved_at_utc=retrieved_at_utc,
                    raw_sha256=raw_sha256,
                    normalized_sha256=write_dataframe_atomic(df, normalized_path),
                    df=df,
                    column_units=WEATHER_FORECAST_COLUMN_UNITS,
                    license_note=OPEN_METEO_LICENSE_NOTE,
                    time_column="valid_time_utc",
                ),
            )
        print(
            f"[weather-forecast] wrote run {run_id} ({len(config.locations)} locations)"
        )


# ---------------------------------------------------------------------------
# Electricity prices (Energy-Charts)
# ---------------------------------------------------------------------------

ENERGY_CHARTS_URL = "https://api.energy-charts.info/price"

# Energy-Charts rejects CY; Cyprus stays without prices rather than pulling in
# a second provider now.
SUPPORTED_MARKETS = {"DE-LU", "FR", "PL", "BG", "IE(SEM)", "NO1"}

PRICE_COLUMN_UNITS = {
    "market_id": "string",
    "delivery_start_utc": "timestamp[UTC]",
    "price_eur_per_mwh": "EUR/MWh",
    "source": "string",
}


def price_request(zone: str, period: Period) -> dict[str, str]:
    """Energy-Charts request covering ``period`` as whole inclusive days.

    Energy-Charts treats ``end`` as an inclusive instant, so asking for
    ``period.end T00:00`` returns only the first hour of the final day and
    leaves the period 23 hours shorter than the weather artifacts, which are
    fetched by inclusive ``start_date``/``end_date``. Requesting the following
    midnight covers the whole period; the extra boundary row is trimmed during
    normalization so consecutive periods stay disjoint.
    """
    return {
        "bzn": zone,
        "start": f"{period.start.isoformat()}T00:00",
        "end": f"{(period.end + timedelta(days=1)).isoformat()}T00:00",
    }


def download_prices(request: dict[str, str]) -> dict[str, Any]:
    zone = request["bzn"]
    if zone not in SUPPORTED_MARKETS:
        raise ValueError(
            f"Unsupported Energy-Charts bidding zone: {zone!r}. "
            f"Supported zones: {sorted(SUPPORTED_MARKETS)}"
        )
    response = _get(ENERGY_CHARTS_URL, request)
    response.raise_for_status()
    return response.json()


def normalize_price_response(
    payload: dict[str, Any], zone: str, end_exclusive: date | None = None
) -> pd.DataFrame:
    """Normalize a price response, preserving native resolution and sign.

    ``end_exclusive`` drops the trailing boundary row that belongs to the next
    period. Resampling and gap filling are benchmark policy and belong in the
    loader.
    """
    unit = payload.get("unit", "")
    if "eur" not in unit.lower() or "mwh" not in unit.lower():
        raise ValueError(f"Unexpected Energy-Charts price unit: {unit!r}")

    df = pd.DataFrame(
        {
            "market_id": zone,
            "delivery_start_utc": pd.to_datetime(
                payload["unix_seconds"], unit="s", utc=True
            ),
            "price_eur_per_mwh": pd.array(payload["price"], dtype="Float64"),
            "source": "energy_charts",
        }
    )

    times = df["delivery_start_utc"]
    if times.duplicated().any():
        raise ValueError(f"Duplicate delivery timestamps in {zone} price response")
    if not times.is_monotonic_increasing:
        raise ValueError(f"Non-increasing delivery timestamps in {zone} price response")

    if end_exclusive is not None:
        boundary = pd.Timestamp(end_exclusive, tz="UTC")
        df = df[df["delivery_start_utc"] < boundary].reset_index(drop=True)
    return df


def run_prices(config: BenchmarkConfig, output_dir: Path, *, force: bool) -> None:
    for location in config.locations:
        if not location.market:
            print(f"[prices] skipping {location.id}: no supported market")
            continue

        for period in config.periods:
            name = f"{location.market.replace('(', '_').replace(')', '')}_{period.id}"
            normalized_path = (
                output_dir / "normalized" / "electricity_prices" / f"{name}.parquet"
            )
            manifest_path = (
                output_dir / "manifests" / "electricity_prices" / f"{name}.json"
            )
            raw_path = output_dir / "raw" / "energy_charts" / f"prices_{name}.json"
            request = price_request(location.market, period)

            if not force and artifact_up_to_date(
                normalized_path, manifest_path, request
            ):
                print(f"[prices] up to date: {normalized_path}")
                continue

            retrieved_at_utc = read_raw_cache(raw_path, request, force=force)
            if retrieved_at_utc:
                print(f"[prices] reusing cached response: {raw_path}")
                payload = json.loads(raw_path.read_text())
                raw_sha256 = sha256_file(raw_path)
            else:
                print(f"[prices] downloading {location.market} ({period.id})")
                payload = download_prices(request)
                retrieved_at_utc = utc_now_iso()
                raw_sha256 = write_bytes_atomic(
                    json.dumps(payload).encode("utf-8"), raw_path
                )
                write_raw_cache(
                    raw_path,
                    source_url=ENERGY_CHARTS_URL,
                    request=request,
                    retrieved_at_utc=retrieved_at_utc,
                )

            df = normalize_price_response(
                payload, location.market, end_exclusive=period.end + timedelta(days=1)
            )
            write_json(
                manifest_path,
                build_manifest(
                    artifact_type="electricity_prices",
                    source_url=ENERGY_CHARTS_URL,
                    request=request,
                    retrieved_at_utc=retrieved_at_utc,
                    raw_sha256=raw_sha256,
                    normalized_sha256=write_dataframe_atomic(df, normalized_path),
                    df=df,
                    column_units=PRICE_COLUMN_UNITS,
                    # Zone-specific: DE-LU/FR/PL are CC BY 4.0, while
                    # BG/IE(SEM)/NO1 are private/internal use only.
                    license_note=payload.get(
                        "license_info", "See Energy-Charts API terms of use."
                    ),
                    time_column="delivery_start_utc",
                ),
            )
            print(f"[prices] wrote {normalized_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

# Forecast runs are excluded from a default run: the full archive is ~20,000
# weighted Open-Meteo calls (multiple days against the free tier) and nothing
# consumes forecasts yet. Ask for them explicitly, and use --max-forecast-runs
# to pull a small subset.
DEFAULT_SOURCES = ["tabula", "weather-reference", "prices"]

SOURCES = {
    "tabula": lambda config, output_dir, args: run_tabula(output_dir, force=args.force),
    "weather-reference": lambda config, output_dir, args: run_weather_reference(
        config, output_dir, force=args.force
    ),
    "weather-forecast": lambda config, output_dir, args: run_weather_forecasts(
        config,
        output_dir,
        force=args.force,
        max_runs=args.max_forecast_runs,
        shard_index=args.forecast_shard_index,
        shard_count=args.forecast_shards,
    ),
    "prices": lambda config, output_dir, args: run_prices(
        config, output_dir, force=args.force
    ),
}


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Download and normalize TABULA building records, Open-Meteo weather, "
            "and Energy-Charts day-ahead electricity prices for the multi-building "
            "benchmark. Writes raw/, normalized/, and manifests/ under --output-dir. "
            "Valid existing artifacts are reused so a long acquisition can be "
            "resumed; never invoked implicitly by I4B simulation code. "
            "A default run covers TABULA, reference weather, and prices; archived "
            "forecast runs are large and must be requested with "
            "--only weather-forecast."
        ),
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path(__file__).parent / "benchmark_source_data.json",
        help="Acquisition configuration: periods, locations, model, markets (default: %(default)s)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        required=True,
        help="Directory to write raw/, normalized/, and manifests/ into",
    )
    parser.add_argument(
        "--only",
        choices=list(SOURCES),
        help="Restrict acquisition to a single source (for development)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Replace existing raw and normalized artifacts instead of reusing them",
    )
    parser.add_argument(
        "--max-forecast-runs",
        type=int,
        metavar="N",
        help="Stop after N forecast runs, for a small subset to develop against",
    )
    parser.add_argument(
        "--forecast-shards",
        type=int,
        default=1,
        help="Split configured forecast initialization times into N disjoint shards",
    )
    parser.add_argument(
        "--forecast-shard-index",
        type=int,
        default=0,
        help="Zero-based forecast shard processed by this invocation",
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    args = build_arg_parser().parse_args(argv)
    if args.forecast_shards < 1:
        raise ConfigError("--forecast-shards must be positive")
    if not 0 <= args.forecast_shard_index < args.forecast_shards:
        raise ConfigError("--forecast-shard-index must be less than --forecast-shards")
    config = load_config(args.config)
    selected = [args.only] if args.only else DEFAULT_SOURCES
    for name in selected:
        SOURCES[name](config, args.output_dir, args)
    if not args.only:
        print("[skip] weather-forecast: run with --only weather-forecast to acquire")


if __name__ == "__main__":
    main(sys.argv[1:])
